# Group 6 
# Maddie Parsons, Seth Gould, Daniel Tsao, David Medina, Jackson Stone, Sarah Gastelum

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Creates a font to bold the headers 
font1 = Font(bold = True)

#create the workbook object
myWorkbook = Workbook()

#load information from excel file
myWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")

#creates individual sheets and removes the default sheet 
outPutWorkbook = Workbook()
outPutWorkbook.create_sheet("Algebra")
outPutWorkbook.create_sheet("Trigonometry")
outPutWorkbook.create_sheet("Geometry")
outPutWorkbook.create_sheet("Calculus")
outPutWorkbook.create_sheet("Statistics")
outPutWorkbook.remove(outPutWorkbook["Sheet"])

# creates the headers for each sheet and sizes the column headers to the length plus 5
for sheet in outPutWorkbook.sheetnames:
    outPutWorkbook[sheet]["A1"].value = "Last Name"
    outPutWorkbook[sheet]["A1"].font = font1
    outPutWorkbook[sheet].column_dimensions["A"].width = len(outPutWorkbook[sheet]["A1"].value) + 5
    outPutWorkbook[sheet]["B1"].value = "First Name"
    outPutWorkbook[sheet]["B1"].font = font1
    outPutWorkbook[sheet].column_dimensions["B"].width = len(outPutWorkbook[sheet]["B1"].value) + 5
    outPutWorkbook[sheet]["C1"].value = "Student ID"
    outPutWorkbook[sheet]["C1"].font = font1
    outPutWorkbook[sheet].column_dimensions["C"].width = len(outPutWorkbook[sheet]["C1"].value) + 5
    outPutWorkbook[sheet]["D1"].value = "Grade"
    outPutWorkbook[sheet]["D1"].font = font1
    outPutWorkbook[sheet].column_dimensions["D"].width = len(outPutWorkbook[sheet]["D1"].value) + 5
    outPutWorkbook[sheet]["F1"].value = "Summary Statistics"
    outPutWorkbook[sheet]["F1"].font = font1
    outPutWorkbook[sheet].column_dimensions["F"].width = len(outPutWorkbook[sheet]["F1"].value) + 5
    outPutWorkbook[sheet]["G1"].value = "Value"
    outPutWorkbook[sheet]["G1"].font = font1
    outPutWorkbook[sheet].column_dimensions["G"].width = len(outPutWorkbook[sheet]["G1"].value) + 5

# iterates through the rows and splits the name, appends the values to the output workbook into the correct sheet
for row in myWorkbook["Grades"].iter_rows(min_row=2, values_only=True):
    subject = row[0]  # Column A
    fullname = row[1]  # Column B
    grade = row[2]     # Column C
    if subject in outPutWorkbook.sheetnames:
            last, first, student_id = fullname.split("_")
    outPutWorkbook[subject].append([last, first, student_id, grade])

# Creates the labels for the stats of the grades
for sheet in outPutWorkbook.sheetnames:
    outPutWorkbook[sheet]["F2"].value = "Highest Grade"
    outPutWorkbook[sheet]["F3"].value = "Lowest Grade"
    outPutWorkbook[sheet]["F4"].value = "Mean Grade"
    outPutWorkbook[sheet]["F5"].value = "Median Grade"
    outPutWorkbook[sheet]["F6"].value = "Number of Students"

# Calcuates the max, min, average, median, and count for each sheet based on the data
for sheet in outPutWorkbook.sheetnames:
    last_row = outPutWorkbook[sheet].max_row
    outPutWorkbook[sheet]["G2"].value = f"=MAX(D2:D{last_row})"
    outPutWorkbook[sheet]["G3"].value = f"=MIN(D2:D{last_row})"
    outPutWorkbook[sheet]["G4"].value = f"=AVERAGE(D2:D{last_row})"
    outPutWorkbook[sheet]["G5"].value = f"=MEDIAN(D2:D{last_row})"
    outPutWorkbook[sheet]["G6"].value = f"=COUNT(D2:D{last_row})"

# adds filters to each sheet 
for sheet in outPutWorkbook.sheetnames: 
    # recalculates the last row to ensure it is correct 
    last_row = outPutWorkbook[sheet].max_row 
    outPutWorkbook[sheet].auto_filter.ref = f"A1:D{last_row}"
    
# Saves the workbook
outPutWorkbook.save(filename = "formatted_grades.xlsx")

# Closes the workbook 
outPutWorkbook.close()